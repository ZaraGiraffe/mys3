# utils.py

import csv
import os
import random
from models import Timetable, Group, Subject, Lecturer, Room
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Constants for time slots (4 periods x 5 days)
TIME_SLOTS: List[Tuple[int, int]] = [(day, period) for day in range(5) for period in range(4)]
DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

def load_data() -> Tuple[List[Group], List[Subject], List[Lecturer], List[Room]]:
    groups: List[Group] = []
    subjects: Dict[str, Subject] = {}
    lecturers_dict: Dict[str, Lecturer] = {}
    lecturers: List[Lecturer] = []
    rooms: List[Room] = []

    # Load subjects
    with open("data/subjects.csv", mode="r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            subject = Subject(
                name=row["name"],
                lecture_hours=int(row["lecture_hours"]),
                practice_hours=int(row["practice_hours"])
            )
            subjects[subject.name] = subject

    # Load groups
    with open("data/groups.csv", mode="r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            subjects_list = [s.strip() for s in row["subjects"].split(",")]
            group = Group(
                group_id=row["group_id"],
                size=int(row["size"]),
                subjects=subjects_list
            )
            groups.append(group)

    # Load lecturers
    with open("data/lecturers.csv", mode="r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            lecturer_id = row["lecturer_id"]
            name = row["name"]
            subject_name = row["subject"].strip()
            available_type = row["available_type"].strip()  # Lecture or Practice

            if lecturer_id not in lecturers_dict:
                lecturer = Lecturer(lecturer_id=lecturer_id, name=name)
                lecturers_dict[lecturer_id] = lecturer
            else:
                lecturer = lecturers_dict[lecturer_id]

            # Add subject and available_type
            if subject_name not in lecturer.subjects:
                lecturer.subjects[subject_name] = []
            lecturer.subjects[subject_name].append(available_type)

    lecturers = list(lecturers_dict.values())

    # Load rooms
    with open("data/rooms.csv", mode="r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            room = Room(
                room_id=row["room_id"],
                capacity=int(row["capacity"])
            )
            rooms.append(room)

    return groups, list(subjects.values()), lecturers, rooms

def check_constraints(timetable: Timetable, groups: List[Group]) -> int:
    violations = 0
    lecturer_schedule: Dict[Tuple[str, Tuple[int, int]], bool] = {}
    room_schedule: Dict[Tuple[str, Tuple[int, int]], bool] = {}
    group_schedule: Dict[Tuple[str, Tuple[int, int]], bool] = {}

    for (group_id, slot), (subject, lecturer, room, session_type) in timetable.schedule.items():
        # Check group schedule
        if group_schedule.get((group_id, slot)):
            violations += 1  # Group is scheduled for more than one class at the same time
        else:
            group_schedule[(group_id, slot)] = True

        # Check lecturer schedule
        if lecturer_schedule.get((lecturer.name, slot)):
            violations += 1  # Lecturer is scheduled for more than one class at the same time
        else:
            lecturer_schedule[(lecturer.name, slot)] = True

        # Check room schedule
        if room_schedule.get((room.room_id, slot)):
            violations += 1  # Room is scheduled for more than one class at the same time
        else:
            room_schedule[(room.room_id, slot)] = True

    return violations

def schedule_to_csv(timetable: Timetable, filename: str = "schedule.csv") -> None:
    with open(filename, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Group ID", "Day", "Period", "Subject", "Lecturer", "Room ID", "Session Type"])

        for (group_id, (day, period)), (subject, lecturer, room, session_type) in timetable.schedule.items():
            writer.writerow([group_id, day, period, subject.name, lecturer.name, room.room_id, session_type])
    print(f"Schedule saved to {filename}")

def save_schedule_to_excel(timetable: Timetable, output_folder: str = "schedules") -> None:
    import os
    from openpyxl import Workbook

    # Create the output directory if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    groups = set(group_id for (group_id, _) in timetable.schedule.keys())

    for group_id in groups:
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Group {group_id}"

        # Set up the header row with days of the week
        for col_num, day in enumerate(DAYS_OF_WEEK, start=2):  # Start from column 2
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = day

        # Set up the first column with period numbers
        for row_num in range(2, 6):  # Periods 1 to 4
            ws[f"A{row_num}"] = f"Period {row_num - 1}"

        # Fill in the schedule
        for (gid, (day, period)), (subject, lecturer, room, session_type) in timetable.schedule.items():
            if gid == group_id:
                # Excel rows and columns are 1-indexed
                col_num = day + 2  # Offset by 2 because column 1 is for periods, and days start from 0
                row_num = period + 2  # Offset by 2 because row 1 is for headers, and periods start from 0
                col_letter = get_column_letter(col_num)
                cell = f"{col_letter}{row_num}"
                ws[cell] = f"{subject.name} : {lecturer.name} : {room.room_id} :: {session_type}"

        # Save the workbook
        filename = os.path.join(output_folder, f"group_{group_id}.xlsx")
        wb.save(filename)
        print(f"Schedule for Group {group_id} saved to {filename}")